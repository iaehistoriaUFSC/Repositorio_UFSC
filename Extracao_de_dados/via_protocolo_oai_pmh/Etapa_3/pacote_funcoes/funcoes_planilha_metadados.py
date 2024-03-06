import time
import os
import re
from unidecode import unidecode
import pandas as pd
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
import openpyxl.utils
import openpyxl
import joblib


def criarDiretorio(caminho : str) -> None:
    if not os.path.exists(caminho):
        os.makedirs(caminho)
        return caminho
    else:
        return caminho
    
def obterDiretorioAtual() -> str:
    """
    Essa função retornará o diretório atual da pasta de trabalho que está executando o 
    script.

    Retorno:
    -------
    - :return: String contendo o caminho total até a pasta de trabalho que está executando 
    este script.
    """
    return os.getcwd()

diretorio_atual = obterDiretorioAtual()

if 'Etapa_3' not in diretorio_atual:
    diretorio_atual = criarDiretorio(os.path.join(diretorio_atual,'Etapa_3'))

caminho_pasta_planilhas = os.path.join(os.path.dirname(diretorio_atual),'Resultados','Planilhas de metadados')
caminho_pasta_dicts = os.path.join(os.path.dirname(diretorio_atual),'Etapa 2','Dicts atualizados')


def formatarNomeArquivoColecao(nome_colecao : str):
    nome_colecao_formatado = re.sub(r'[^\w]', '_', unidecode(nome_colecao))
    nome_colecao_formatado = nome_colecao_formatado.replace('__','_')
    if nome_colecao_formatado.endswith('_'):
        nome_colecao_formatado = nome_colecao_formatado[:len(nome_colecao_formatado)-1]
    nome_colecao_formatado = nome_colecao_formatado.replace('Programa_de_Pos_Graduacao_em_','')
    return nome_colecao_formatado


def criarGeradorDicts(lista_de_dicionarios : list):
    """
    Cria um gerador de dicionários das coleções a partir de uma lista com os
    caminhos para cada dicionário. Nessa lista "geradora" os elementos serão
    da seguinte forma: ('caminho até o dicionário',dicionário em questão)
    Dessa forma tem-se acesso ao nome do dicionário e ao dicionário em si.
    Ou seja, a lista "geradora" terá a seguinte tipagem: [(str,dict)].
    Observação: Como é um objeto gerador não dá para acessar diretamente os
    elementos pelas suas posições, somente acessá-los um a um, iniciando do
    primeiro (sem conseguir retroceder).

    :param lista_de_dicionarios: Lista contendo os caminhos de todos os
    dicionários que serão contemplados nesta execução.
    :return: Objeto gerador de tuplas com os caminhos dos dicionários
    e os dicionários carregados, respectivamente.
    """
    for arquivo_dic in lista_de_dicionarios:
        if os.path.basename(arquivo_dic) != 'Ecossistemas_Agricolas_e_Naturais.joblib':
            yield (arquivo_dic.replace('.joblib',''),joblib.load(arquivo_dic))

def criarExcel(dicionarios,
               caminho_pasta_planilhas : str,
               nome_da_planilha : str) -> tuple[bool, str]:
    """
    Cria um Excel (planilha de metadados) com as abas sendo as coleções,
    cada linha um trabalho e os metadados distribuídos nas suas respectivas
    colunas.

    :param dicionarios: Gerador de dicionários (como se fosse uma "lista"),
    onde cada elemento representa um dicionário no formato: "('caminho para
    o dicionário', dicionário propriamente dito)".
    :param caminho_pasta_planilhas: String que representa o caminho até a pasta
    que é armazenada a planilha.
    :param nome_da_planilha: String referente ao nome que será dado ao arquivo
    Excel que representa a planilha de metadados gerais (que contém todas as
    coleções).
    :return: Retorna uma tupla com o status do processo (True se tudo ocorreu
    bem e False se algo saiu do esperado) e uma mensagem de erro caso o status
    seja False. Caso status seja True retorna junto a mensagem "''".
    """
    # :param fill_laranja_suave: Responsável por fazer o preenchimento das
    # células pertinentes (língua diferente de português e ano menor que 2003)
    # na cor laranja suave.
    # :param fill_vermelho_suave: Responsável por fazer o preenchimento das
    # células pertinentes (com valor igual a "N.I.") na cor vermelho suave.
    # :param link_style: Responsável por modificar o estilo que os links
    # aparecerão nas células da planilha (uma espécie de máscara).
    # :param borda: Responsável por modificar o estilo das bordas das células
    # da planilha, deixando-as cinzas claras.

    fill_laranja_suave = PatternFill(start_color="ffab57", end_color="ffab57", fill_type="solid")
    fill_vermelho_suave = PatternFill(start_color="ff5757", end_color="ff5757", fill_type="solid")
    link_style = Font(underline='single', color='0563C1', size=10)
    cor_cinza_claro = 'D3D3D3'
    borda = Border(left=Side(style='thin', color=cor_cinza_claro),
                  right=Side(style='thin', color=cor_cinza_claro),
                  top=Side(style='thin', color=cor_cinza_claro),
                  bottom=Side(style='thin', color=cor_cinza_claro))
    try:
        with pd.ExcelWriter(os.path.join(caminho_pasta_planilhas,nome_da_planilha), engine='openpyxl') as writer:
            print('\nCriando planilha:',nome_da_planilha)
            for dicionario in dicionarios:
                try:
                    nome_da_colecao_atual = dicionario[0]
                    print('\nPassando pela coleção:',os.path.basename(nome_da_colecao_atual))
                    df = pd.DataFrame(dicionario[1])
                    nome_sheet = os.path.splitext(os.path.basename(nome_da_colecao_atual))[0].replace('Mestrado_Profissional','MP').replace('Universitaria','UNIV').replace('Desenvolvimento','DESENV').replace('Biotecnologia','BIOTEC').replace('Biologia','BIO').replace('Administracao','ADM').replace('Engenharia','ENG').replace('Historia','HST').replace('__','_').replace('_de_','_')

                    df.to_excel(writer, sheet_name=nome_sheet, index=False)

                    # Acesse a folha Excel recém-criada
                    worksheet = writer.sheets[nome_sheet]

                    # Defina o estilo de alinhamento para as colunas desejadas
                    for col_num, col_name in enumerate(df.columns, 1):
                        col_letter = openpyxl.utils.get_column_letter(col_num)

                        # Centralize todas as células na coluna 'Língua'
                        if col_name == 'Língua':
                            worksheet.column_dimensions[col_letter].width = 11
                            for row_num in range(2, len(df) + 2):  # Inicia do segundo row (1-indexed)
                                cell = worksheet['{}{}'.format(col_letter, row_num)]
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                if (str(cell.value).strip().lower() != 'por') and (str(cell.value).strip().lower() != 'pt_br'):
                                    cell.fill = fill_laranja_suave

                        # Centralize todas as células nas colunas 'Ano repositório' e 'Ano descrição'
                        elif col_name in ['Ano repositório', 'Ano descrição','Ano capa']:
                            for row_num in range(2, len(df) + 2):  # Inicia do segundo row (1-indexed)
                                cell = worksheet['{}{}'.format(col_letter, row_num)]
                                cell.alignment = Alignment(horizontal='center', vertical='center')

                                # Pinte de laranja suave se o valor for menor que 2003
                                cell_value = str(df.at[row_num - 2, col_name])
                                if cell_value.isdigit():
                                    if int(cell_value) < 2003:
                                        cell.fill = fill_laranja_suave
                            worksheet.column_dimensions[col_letter].width = 21

                        elif col_name in ['Link PDF','Link página']:
                            worksheet.column_dimensions[col_letter].width = 18
                            for row_num in range(2, len(df) + 2):  # Inicia do segundo row (1-indexed)
                                cell = worksheet['{}{}'.format(col_letter, row_num)]
                                # Centralizar e ativar a quebra de texto
                                cell.alignment = Alignment(shrink_to_fit=True)

                                # Mudando o valor do link para algo pré-setado (para obter o valor depois, ao carregá-la tem que usar "sheet['A2'].hyperlink.target" ou "cel.hyperlink.target")
                                # Exemplo disso seria:
                                ############################################################################################################################################
                                # import pandas as pd
                                # from openpyxl import load_workbook
                                #
                                # # Ler o arquivo Excel usando Pandas
                                # df = pd.read_excel('exemplo_links_clicaveis.xlsx')
                                #
                                # # Acessar o valor e o hyperlink de uma célula específica (por exemplo, A2)
                                # linha = 1  # Lembrando que o Pandas usa índice 0 para as linhas
                                # coluna = 0  # Índice 0 para a coluna A
                                # valor_celula = df.iloc[linha, coluna]
                                # hyperlink_celula = df.iloc[linha, coluna].hyperlink
                                #
                                # print(f"Valor da célula A2: {valor_celula}")
                                # print(f"Hyperlink da célula A2: {hyperlink_celula}")
                                ############################################################################################################################################

                                cell.font = link_style
                                if cell.value != 'N.I.':
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                                    if col_name == 'Link PDF':
                                        # Adicionar um link clicável na célula A2 (substitua pelo seu próprio link)
                                        cell.hyperlink = cell.value
                                        cell.value = "PDF"
                                    else:
                                        worksheet[col_letter+str(row_num)].hyperlink = cell.value
                                        worksheet[col_letter+str(row_num)].value = "Página"


                        elif col_name == 'Descrição':
                            worksheet.column_dimensions[col_letter].width = 15
                        elif col_name == 'Texto extraído?':
                            worksheet.column_dimensions[col_letter].width = 18
                            for row_num in range(2, len(df) + 2):
                                cell = worksheet['{}{}'.format(col_letter, row_num)]
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            worksheet.column_dimensions[col_letter].width = 13

                    for r_idx, row in df.iterrows():
                        for c_idx, value in enumerate(row, start=1):
                            cell = worksheet.cell(row=r_idx + 2, column=c_idx)
                            cell.border = borda
                            if value == 'N.I.' or value == 'NÃO':
                                cell.fill = fill_vermelho_suave
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                cell.font = Font(bold=True, size=10, name='Arial')
                            else:
                                cell.font = Font(size=10, name='Arial')
                    print(f'Coleção: {nome_sheet} foi adicionada com sucesso!')
                except Exception as e:
                    erro = f'Erro "{e.__class__.__name__}": {str(e)}'
                    print(f'! Problema ao gerar planilha --> {erro}')

        return True, ''
    except Exception as e:
        erro = f'Erro "{e.__class__.__name__}": {str(e)}'
        return False, erro

def gerarPlanilhaDeMetadados(lista_de_dicionarios : list,
                             caminho_pasta_planilhas : str,
                             nome_da_planilha : str) -> tuple[bool,str]:
    """
    Função que organiza a chamada das funções responsáveis pela criação da
    planilha de metadados gerais (com todas as coleções). Foi criada a fim de
    organizar a função main() de tal forma que tenha um processo específico para
    a geração da planilha de metadados e outro (outra função) específico para
    geração do relatório em PDF.

    :param lista_de_dicionarios: Lista contendo os caminhos de todos os
    dicionários que serão contemplados nesta execução
    :param caminho_pasta_planilhas: String que representa o caminho até a pasta
    que é armazenada a planilha.
    :param nome_da_planilha: String referente ao nome que será dado ao arquivo
    Excel que representa a planilha de metadados gerais (que contém todas as
    coleções).
    :return: Retorna uma tupla com o status do processo (True se tudo ocorreu
    bem e False se algo saiu do esperado) e uma mensagem de erro caso o status
    seja False. Caso status seja True retorna junto a mensagem "''".
    """
    try:
        dicionarios = criarGeradorDicts(lista_de_dicionarios=lista_de_dicionarios) # Armazenando dicts nessa variável com gerador


        status_criar_excel, msg_criar_excel = criarExcel(dicionarios=dicionarios,
                                                        caminho_pasta_planilhas=caminho_pasta_planilhas,
                                                        nome_da_planilha=nome_da_planilha)

        if status_criar_excel:
            return True, ''
        else:
            return False, msg_criar_excel + '\nFunção criarExcel()'
    except Exception as e:
        erro = f'Erro "{e.__class__.__name__}": {str(e)}'
        return False, erro + '\nFunção gerarPlanilhaDeMetadados()'

